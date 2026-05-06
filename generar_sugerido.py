"""
generar_sugerido.py
Walmart Stationery Predictive Analytics - Purchase Order Suggestion Generator

Usage:
    python generar_sugerido.py input.xlsx --lead-time 14 --output sugerido_compra.xlsx
    python generar_sugerido.py input.csv --lead-time 7 --whpck 6 --vndpk 12 --output out.xlsx
"""

import argparse
import csv
import math
import os
import sys
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl is required. Install it with: pip install openpyxl")

# ---------------------------------------------------------------------------
# Column alias maps (lowercase key -> canonical name)
# ---------------------------------------------------------------------------
ALIAS = {
    "storenb":    "storeNbr",  "stornbr": "storeNbr",
    "storenbr":   "storeNbr",  "tienda":  "storeNbr",
    "store_nbr":  "storeNbr",  "store":   "storeNbr",

    "storename":  "storeName", "nombre_tienda": "storeName",
    "store_name": "storeName",

    "formato":    "formato",   "format":  "formato",

    "whsenbr":    "whseNbr",   "cedis":   "whseNbr",
    "wh":         "whseNbr",   "warehouse": "whseNbr",
    "wh_nbr":     "whseNbr",   "bodeganbr": "whseNbr",

    "itemnbr":    "itemNbr",   "articulo": "itemNbr",
    "item_nbr":   "itemNbr",   "sku":      "itemNbr",
    "item":       "itemNbr",

    "itemdesc":   "itemDesc",  "descripcion": "itemDesc",
    "item_desc":  "itemDesc",  "desc":        "itemDesc",
    "description": "itemDesc",

    "onhand":     "onHand",    "inv_mano": "onHand",
    "oh":         "onHand",    "on_hand":  "onHand",

    "intransit":  "inTransit", "en_transito": "inTransit",
    "in_transit": "inTransit", "transito":    "inTransit",
    "transit":    "inTransit",

    "inwhse":     "inWhse",    "en_bodega": "inWhse",
    "in_whse":    "inWhse",    "bodega":    "inWhse",
    "in_warehouse": "inWhse",

    "onorder":    "onOrder",   "en_orden": "onOrder",
    "on_order":   "onOrder",   "orden":    "onOrder",
    "order":      "onOrder",

    "promdia":    "promDia",   "venta_diaria": "promDia",
    "prom_dia":   "promDia",   "avg_daily":    "promDia",
    "sales_avg":  "promDia",   "avg_sales":    "promDia",
    "prom_ventas": "promDia",

    "pos2sem":    "pos2sem",   "pos_2sem":    "pos2sem",
    "ventas_2sem": "pos2sem",  "pos_reciente": "pos2sem",
    "pos_2weeks": "pos2sem",

    "whpck":      "whpck",     "wh_pack":  "whpck",
    "whpack":     "whpck",     "wh_pck":   "whpck",

    "vndpk":      "vndpk",     "vnd_pack": "vndpk",
    "vndpack":    "vndpk",     "vendor_pack": "vndpk",
    "vnd_pk":     "vndpk",

    "leadtime":   "leadTime",  "lead_time": "leadTime",
    "plazo":      "leadTime",  "lead":      "leadTime",
}

REQUIRED_COLS = ["storeNbr", "itemNbr", "promDia"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    return int(safe_float(val, default))


def normalize_header(h):
    """Return canonical name for a raw header string, or None if unknown."""
    key = h.strip().lower().replace(" ", "_").replace("-", "_")
    return ALIAS.get(key)


def map_headers(raw_headers):
    """
    Map raw column headers to canonical names.
    Returns (canonical_map, unmapped) where canonical_map = {canonical: index}.
    """
    canonical_map = {}
    unmapped = []
    for i, h in enumerate(raw_headers):
        canon = normalize_header(h)
        if canon and canon not in canonical_map:
            canonical_map[canon] = i
        elif not canon:
            unmapped.append(h)
    return canonical_map, unmapped


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def read_csv(path):
    """Return (headers, rows) from CSV file."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        # Sniff delimiter
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        headers = next(reader)
        rows = list(reader)
    return headers, rows


def read_excel(path):
    """Return (headers, rows) from first sheet of Excel file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows = []
    for row in rows_iter:
        rows.append([c for c in row])
    wb.close()
    return headers, rows


def load_input(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        return read_excel(path)
    elif ext in (".csv", ".txt", ".tsv"):
        return read_csv(path)
    else:
        # Try CSV as default
        return read_csv(path)


# ---------------------------------------------------------------------------
# Core calculation
# ---------------------------------------------------------------------------

def calculate_row(rec, default_lead, default_whpck, default_vndpk):
    """
    rec: dict of canonical_name -> raw value
    Returns enriched dict with calculated fields, or None to skip.
    """
    prom_dia = safe_float(rec.get("promDia"))
    if prom_dia <= 0:
        return None  # Cannot calculate without positive average sales

    on_hand    = safe_float(rec.get("onHand", 0))
    in_transit = safe_float(rec.get("inTransit", 0))
    in_whse    = safe_float(rec.get("inWhse", 0))
    on_order   = safe_float(rec.get("onOrder", 0))

    lead_time  = safe_float(rec.get("leadTime", default_lead), default_lead)
    whpck      = safe_float(rec.get("whpck", default_whpck), default_whpck)
    vndpk      = safe_float(rec.get("vndpk", default_vndpk), default_vndpk)

    if whpck <= 0:
        whpck = 1
    if vndpk <= 0:
        vndpk = whpck

    inv_total       = on_hand + in_transit + in_whse + on_order
    cobertura_actual = inv_total / prom_dia
    falta_lt        = max(0.0, prom_dia * lead_time - inv_total)
    sugerido_whpck  = math.ceil(falta_lt / whpck) * whpck
    sugerido_vndpk  = math.ceil(sugerido_whpck / vndpk) * vndpk

    pos2sem_raw = rec.get("pos2sem")
    has_pos2sem = pos2sem_raw is not None and pos2sem_raw != ""
    pos2sem     = safe_float(pos2sem_raw) if has_pos2sem else None

    incremento = False
    if has_pos2sem and pos2sem is not None:
        incremento = (pos2sem / (prom_dia * 14)) > 1.20

    proximo_agotarse = cobertura_actual < lead_time

    return {
        **rec,
        "_on_hand":          on_hand,
        "_in_transit":       in_transit,
        "_in_whse":          in_whse,
        "_on_order":         on_order,
        "_prom_dia":         prom_dia,
        "_lead_time":        lead_time,
        "_whpck":            whpck,
        "_vndpk":            vndpk,
        "_inv_total":        inv_total,
        "_cobertura_actual": cobertura_actual,
        "_falta_lt":         falta_lt,
        "_sugerido_whpck":   sugerido_whpck,
        "_sugerido_vndpk":   sugerido_vndpk,
        "_incremento":       incremento,
        "_proximo_agotarse": proximo_agotarse,
    }


# ---------------------------------------------------------------------------
# Excel output helpers
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
ROW_FONT      = Font(size=10)
ALT_FILL      = PatternFill("solid", fgColor="DCE6F1")
RED_FILL      = PatternFill("solid", fgColor="FF4C4C")
ORANGE_FILL   = PatternFill("solid", fgColor="FF9933")
YELLOW_FILL   = PatternFill("solid", fgColor="FFEB9C")
THIN_BORDER   = Border(
    left=Side(style="thin", color="B8CCE4"),
    right=Side(style="thin", color="B8CCE4"),
    top=Side(style="thin", color="B8CCE4"),
    bottom=Side(style="thin", color="B8CCE4"),
)


def style_header_row(ws, row_num, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_w=8, max_w=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def write_value(cell, val, fmt=None):
    cell.value = val
    cell.font  = ROW_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if fmt:
        cell.number_format = fmt


def freeze_top_row(ws):
    ws.freeze_panes = ws["A2"]


# ---------------------------------------------------------------------------
# Sheet 1: Detalle Tda-Art
# ---------------------------------------------------------------------------
DETAIL_HEADERS = [
    "storeNbr", "storeName", "formato", "whseNbr",
    "itemNbr", "itemDesc",
    "onHand", "inTransit", "inWhse", "onOrder",
    "promDia", "pos2sem",
    "leadTime", "whpck", "vndpk",
    "inv_total", "cobertura_actual", "falta_lt",
    "sugerido_whpck", "sugerido_vndpk",
    "flag_incremento", "flag_agotarse",
]

DETAIL_HEADER_LABELS = {
    "storeNbr":       "Tienda #",
    "storeName":      "Nombre Tienda",
    "formato":        "Formato",
    "whseNbr":        "CEDIS #",
    "itemNbr":        "Artículo #",
    "itemDesc":       "Descripción",
    "onHand":         "En Mano",
    "inTransit":      "En Tránsito",
    "inWhse":         "En Bodega",
    "onOrder":        "En Orden",
    "promDia":        "Vta Diaria Prom",
    "pos2sem":        "POS 2 Sem",
    "leadTime":       "Lead Time (días)",
    "whpck":          "WH Pack",
    "vndpk":          "Vnd Pack",
    "inv_total":      "Inv Total",
    "cobertura_actual": "Cobertura (días)",
    "falta_lt":       "Falta Lead Time",
    "sugerido_whpck": "Sugerido WH Pack",
    "sugerido_vndpk": "Sugerido Vnd Pack",
    "flag_incremento": "Incremento +20%",
    "flag_agotarse":  "Próx Agotarse",
}

NUM_FMT   = "#,##0.00"
INT_FMT   = "#,##0"
FLOAT_FMT = "#,##0.0"


def write_sheet1(wb, data):
    ws = wb.create_sheet("Detalle Tda-Art")
    headers = [DETAIL_HEADER_LABELS.get(h, h) for h in DETAIL_HEADERS]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 35

    row_n = 2
    for rec in data:
        incr  = rec["_incremento"]
        agot  = rec["_proximo_agotarse"]

        if agot and incr:
            row_fill = RED_FILL
        elif agot:
            row_fill = ORANGE_FILL
        elif incr:
            row_fill = YELLOW_FILL
        else:
            row_fill = ALT_FILL if row_n % 2 == 0 else None

        values = [
            rec.get("storeNbr", ""),
            rec.get("storeName", ""),
            rec.get("formato", ""),
            rec.get("whseNbr", ""),
            rec.get("itemNbr", ""),
            rec.get("itemDesc", ""),
            rec["_on_hand"],
            rec["_in_transit"],
            rec["_in_whse"],
            rec["_on_order"],
            rec["_prom_dia"],
            safe_float(rec.get("pos2sem")) if rec.get("pos2sem") not in (None, "") else "",
            rec["_lead_time"],
            rec["_whpck"],
            rec["_vndpk"],
            rec["_inv_total"],
            rec["_cobertura_actual"],
            rec["_falta_lt"],
            rec["_sugerido_whpck"],
            rec["_sugerido_vndpk"],
            "✓" if incr else "—",
            "✓" if agot else "—",
        ]
        fmts = [
            None, None, None, None, None, None,
            INT_FMT, INT_FMT, INT_FMT, INT_FMT,
            NUM_FMT, INT_FMT,
            INT_FMT, INT_FMT, INT_FMT,
            INT_FMT, FLOAT_FMT, NUM_FMT,
            INT_FMT, INT_FMT,
            None, None,
        ]

        for col_i, (val, fmt) in enumerate(zip(values, fmts), 1):
            cell = ws.cell(row=row_n, column=col_i)
            write_value(cell, val, fmt)
            if row_fill:
                cell.fill = row_fill

        row_n += 1

    freeze_top_row(ws)
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Sugerido x CEDIS-Formato
# ---------------------------------------------------------------------------
AGG_HEADERS = [
    "whseNbr", "formato", "itemNbr", "itemDesc",
    "sum_sugerido_vndpk", "n_tiendas", "vndpk",
]
AGG_LABELS = {
    "whseNbr":            "CEDIS #",
    "formato":            "Formato",
    "itemNbr":            "Artículo #",
    "itemDesc":           "Descripción",
    "sum_sugerido_vndpk": "Total Unid Sugeridas",
    "n_tiendas":          "# Tiendas",
    "vndpk":              "Vnd Pack",
}


def write_sheet2(wb, data):
    # Aggregate: key = (whseNbr, formato, itemNbr)
    agg = {}
    for rec in data:
        key = (
            str(rec.get("whseNbr", "")),
            str(rec.get("formato", "")),
            str(rec.get("itemNbr", "")),
        )
        if key not in agg:
            agg[key] = {
                "whseNbr": rec.get("whseNbr", ""),
                "formato":  rec.get("formato", ""),
                "itemNbr":  rec.get("itemNbr", ""),
                "itemDesc": rec.get("itemDesc", ""),
                "sum_sugerido_vndpk": 0,
                "n_tiendas": 0,
                "vndpk":    rec["_vndpk"],
            }
        agg[key]["sum_sugerido_vndpk"] += rec["_sugerido_vndpk"]
        agg[key]["n_tiendas"] += 1

    rows = sorted(
        agg.values(),
        key=lambda r: (str(r["whseNbr"]), str(r["formato"]), -r["sum_sugerido_vndpk"])
    )

    ws = wb.create_sheet("Sugerido x CEDIS-Formato")
    headers = [AGG_LABELS.get(h, h) for h in AGG_HEADERS]
    for col_i, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 35

    for row_n, rec in enumerate(rows, 2):
        row_fill = ALT_FILL if row_n % 2 == 0 else None
        values = [
            rec["whseNbr"], rec["formato"], rec["itemNbr"], rec["itemDesc"],
            rec["sum_sugerido_vndpk"], rec["n_tiendas"], rec["vndpk"],
        ]
        fmts = [None, None, None, None, INT_FMT, INT_FMT, INT_FMT]
        for col_i, (val, fmt) in enumerate(zip(values, fmts), 1):
            cell = ws.cell(row=row_n, column=col_i)
            write_value(cell, val, fmt)
            if row_fill:
                cell.fill = row_fill

    freeze_top_row(ws)
    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions
    return ws, rows


# ---------------------------------------------------------------------------
# Sheet 3: OC Walmart
# ---------------------------------------------------------------------------
OC_HEADERS = ["Item#", "Description", "CEDIS", "Formato", "Qty_vndpk", "Qty_units", "Pack_size"]


def write_sheet3(wb, agg_rows):
    ws = wb.create_sheet("OC Walmart")
    for col_i, h in enumerate(OC_HEADERS, 1):
        ws.cell(row=1, column=col_i, value=h)
    style_header_row(ws, 1, len(OC_HEADERS))
    ws.row_dimensions[1].height = 30

    for row_n, rec in enumerate(agg_rows, 2):
        vndpk     = rec["vndpk"]
        qty_vndpk = rec["sum_sugerido_vndpk"]
        qty_units  = qty_vndpk  # already in units (sugerido_vndpk is in units, multiples of vndpk)
        row_fill  = ALT_FILL if row_n % 2 == 0 else None
        values = [
            rec["itemNbr"], rec["itemDesc"], rec["whseNbr"], rec["formato"],
            int(qty_vndpk / vndpk) if vndpk else qty_vndpk,
            int(qty_units),
            int(vndpk),
        ]
        fmts = [None, None, None, None, INT_FMT, INT_FMT, INT_FMT]
        for col_i, (val, fmt) in enumerate(zip(values, fmts), 1):
            cell = ws.cell(row=row_n, column=col_i)
            write_value(cell, val, fmt)
            if row_fill:
                cell.fill = row_fill

    freeze_top_row(ws)
    auto_width(ws)
    return ws


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def print_summary(detail_data, agg_rows):
    total_items   = len(set(str(r.get("itemNbr", "")) for r in detail_data))
    total_units   = sum(r["_sugerido_vndpk"] for r in detail_data)
    total_stores  = len(set(str(r.get("storeNbr", "")) for r in detail_data))

    print("\n" + "=" * 60)
    print("  RESUMEN SUGERIDO DE COMPRA")
    print("=" * 60)
    print(f"  Tiendas con sugerido:   {total_stores:>8,}")
    print(f"  Artículos únicos:       {total_items:>8,}")
    print(f"  Total unidades suger.:  {total_units:>8,.0f}")
    print("-" * 60)

    by_formato = defaultdict(lambda: {"units": 0, "rows": 0, "items": set()})
    for rec in detail_data:
        fmt = str(rec.get("formato", "SIN FORMATO"))
        by_formato[fmt]["units"] += rec["_sugerido_vndpk"]
        by_formato[fmt]["rows"]  += 1
        by_formato[fmt]["items"].add(str(rec.get("itemNbr", "")))

    print(f"  {'Formato':<20} {'Artículos':>10} {'Renglones':>10} {'Unidades':>12}")
    print("  " + "-" * 56)
    for fmt in sorted(by_formato):
        d = by_formato[fmt]
        print(f"  {fmt:<20} {len(d['items']):>10,} {d['rows']:>10,} {d['units']:>12,.0f}")

    print("=" * 60)
    print(f"  Combinaciones CEDIS-Formato-Art en OC: {len(agg_rows):,}")
    print("=" * 60 + "\n")


# ---------------------------------------------------------------------------
# Prompt helpers for missing pack sizes
# ---------------------------------------------------------------------------

def prompt_pack(name, arg_val):
    if arg_val is not None and arg_val > 0:
        return arg_val
    while True:
        val = input(f"  Ingresa {name} (empaque por default, > 0): ").strip()
        try:
            v = float(val)
            if v > 0:
                return v
            print("  Debe ser mayor a 0.")
        except ValueError:
            print("  Valor inválido.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Genera sugerido de compra para papelería Walmart."
    )
    parser.add_argument("input", help="Archivo de entrada (.xlsx, .csv, .tsv)")
    parser.add_argument(
        "--lead-time", type=float, default=14,
        metavar="DAYS",
        help="Lead time en días (default: 14). Se ignora si la columna existe en el archivo."
    )
    parser.add_argument(
        "--whpck", type=float, default=None,
        metavar="N",
        help="WH Pack default si no está en el archivo."
    )
    parser.add_argument(
        "--vndpk", type=float, default=None,
        metavar="N",
        help="Vendor Pack default si no está en el archivo."
    )
    parser.add_argument(
        "--output", "-o", default=None,
        metavar="FILE",
        help="Nombre del archivo Excel de salida (default: sugerido_compra.xlsx)."
    )
    parser.add_argument(
        "--no-prompt", action="store_true",
        help="No preguntar de forma interactiva si faltan datos; usar 1 como default para packs."
    )
    return parser.parse_args()


def main():
    args = parse_args()

    if not os.path.exists(args.input):
        sys.exit(f"ERROR: No se encontró el archivo '{args.input}'")

    print(f"\nLeyendo archivo: {args.input}")
    raw_headers, raw_rows = load_input(args.input)
    print(f"  Columnas encontradas: {len(raw_headers)}")
    print(f"  Renglones de datos:   {len(raw_rows)}")

    col_map, unmapped = map_headers(raw_headers)
    if unmapped:
        print(f"  Columnas no reconocidas (ignoradas): {unmapped}")

    # Check required columns
    missing_required = [c for c in REQUIRED_COLS if c not in col_map]
    if missing_required:
        print(f"\nERROR: Faltan columnas requeridas: {missing_required}")
        print(f"  Columnas detectadas: {list(col_map.keys())}")
        print(f"  Columnas raw:        {raw_headers}")
        sys.exit(1)

    has_whpck  = "whpck"    in col_map
    has_vndpk  = "vndpk"   in col_map
    has_lead   = "leadTime" in col_map

    # Resolve pack defaults
    if not has_whpck:
        if args.no_prompt:
            default_whpck = args.whpck if args.whpck else 1.0
        else:
            print("\nLa columna de WH Pack (whpck) no se encontró en el archivo.")
            default_whpck = prompt_pack("WH Pack (whpck)", args.whpck)
    else:
        default_whpck = args.whpck if args.whpck else 1.0

    if not has_vndpk:
        if args.no_prompt:
            default_vndpk = args.vndpk if args.vndpk else default_whpck
        else:
            print("\nLa columna de Vendor Pack (vndpk) no se encontró en el archivo.")
            default_vndpk = prompt_pack("Vendor Pack (vndpk)", args.vndpk)
    else:
        default_vndpk = args.vndpk if args.vndpk else 1.0

    default_lead = args.lead_time if not has_lead else args.lead_time

    print("\nCalculando sugeridos...")

    processed = []
    skipped   = 0

    for raw_row in raw_rows:
        # Build canonical record
        rec = {}
        for canon, idx in col_map.items():
            val = raw_row[idx] if idx < len(raw_row) else None
            rec[canon] = val

        result = calculate_row(rec, default_lead, default_whpck, default_vndpk)
        if result is None:
            skipped += 1
            continue
        if result["_sugerido_vndpk"] <= 0:
            continue
        processed.append(result)

    if skipped:
        print(f"  Renglones omitidos (promDia = 0 o inválido): {skipped}")

    # Sort: formato ASC, whseNbr ASC, itemNbr ASC, storeNbr ASC
    processed.sort(key=lambda r: (
        str(r.get("formato", "")),
        str(r.get("whseNbr", "")),
        str(r.get("itemNbr", "")),
        str(r.get("storeNbr", "")),
    ))

    print(f"  Renglones con sugerido > 0: {len(processed)}")

    if not processed:
        print("\nNo hay sugeridos que generar. Verifica los datos de entrada.")
        sys.exit(0)

    # Build workbook
    out_path = args.output or "sugerido_compra.xlsx"
    print(f"\nGenerando archivo Excel: {out_path}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_sheet1(wb, processed)
    print("  Hoja 1 'Detalle Tda-Art' creada.")

    _, agg_rows = write_sheet2(wb, processed)
    print("  Hoja 2 'Sugerido x CEDIS-Formato' creada.")

    write_sheet3(wb, agg_rows)
    print("  Hoja 3 'OC Walmart' creada.")

    wb.save(out_path)
    print(f"  Archivo guardado: {out_path}")

    print_summary(processed, agg_rows)


if __name__ == "__main__":
    main()
